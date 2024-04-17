import tkinter as tk
from tkinter import filedialog
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Alignment
from tkinter import ttk
import os
from PIL import ExifTags

def upload_image():
    filepath = filedialog.askopenfilename(
        title="Selecione a imagem", 
        filetypes=(("Arquivos de imagem", ".jpg;.jpeg;.png;.bmp"), ("Todos os arquivos", ".")))
    if filepath:
        image = Image.open(filepath)
        metadata = image.info
        exif_data = image._getexif()
        if exif_data is not None:
            for tag, value in exif_data.items():
                tagname = ExifTags.TAGS.get(tag, tag)
                metadata[tagname] = value
        else:
            print("Esta imagem não contém metadados EXIF.")
        display_metadata(metadata)
        save_metadata(metadata)

def get_image_metadata(filepath):
    metadata = {}
    with open(filepath, 'rb') as f:
        metadata.update(dict(Image.open(f).info.items()))
    return metadata

def display_metadata(metadata):
    for widget in frame_metadata.winfo_children():
        widget.destroy()
    treeview_metadata = ttk.Treeview(frame_metadata, columns=("Key", "Value"), show="headings")
    treeview_metadata.heading("Key", text="Key")
    treeview_metadata.heading("Value", text="Value")
    for key, value in metadata.items():
        treeview_metadata.insert("", "end", values=(key, value))
    treeview_metadata.pack(fill="both", expand=True)

def save_metadata(metadata):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Metadados"
    
    # Escrevendo cabeçalhos
    headers = ["Chave", "Valor"]
    sheet.append(headers)
    
    # Escrevendo metadados
    for key, value in metadata.items():
        sheet.append([str(key), str(value)])
    
    # Ajustando largura das colunas
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
    
    # Alinhando células
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Salvando arquivo
    workbook.save("metadados.xlsx")
    label_status.config(text="Metadados salvos com sucesso!")

# Configuração da interface
root = tk.Tk()
root.title("Obter Metadados de Imagem")

label_instruction = tk.Label(root, text="Clique no botão abaixo para selecionar uma imagem:")
label_instruction.pack(pady=10)

button_upload = tk.Button(root, text="Upload de Imagem", command=upload_image)
button_upload.pack(pady=5)

frame_metadata = tk.Frame(root)
frame_metadata.pack(pady=10, fill="both", expand=True)

label_status = tk.Label(root, text="")
label_status.pack(pady=5)

root.mainloop()
